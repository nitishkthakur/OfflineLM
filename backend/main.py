import os
import json
import base64
import uuid
import re
from pathlib import Path
from typing import Optional, AsyncGenerator
from datetime import datetime

import httpx

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from dotenv import load_dotenv

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfgen import canvas as _rl_canvas

from tools import make_tools
import agents as agent_registry
from sandbox import SandboxRegistry, ensure_sandbox_venv, bwrap_path

load_dotenv()

app = FastAPI(title="Jarvis API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).parent.parent
CONFIG_PATH = BASE_DIR / "config.json"
ARTIFACTS_DIR = BASE_DIR / "artifacts"
UPLOADS_DIR = BASE_DIR / "uploads"
EXPORTS_DIR = BASE_DIR / "exports"

ARTIFACTS_DIR.mkdir(exist_ok=True)
UPLOADS_DIR.mkdir(exist_ok=True)
EXPORTS_DIR.mkdir(exist_ok=True)

conversations: dict = {}

# Sandbox registry (lazily initialized on first use when code execution is enabled).
_sandbox_registry: Optional[SandboxRegistry] = None
_sandbox_reaper_task = None


def _get_sandbox_registry() -> Optional[SandboxRegistry]:
    """Return the shared sandbox registry, initializing it on first call.

    Returns None if code_execution is disabled in config.
    """
    global _sandbox_registry
    cfg = load_config().get("code_execution", {})
    if not cfg.get("enabled"):
        return None
    if _sandbox_registry is not None:
        return _sandbox_registry

    venv_raw = cfg.get("sandbox_venv", ".sandbox-venv")
    venv_dir = Path(venv_raw)
    if not venv_dir.is_absolute():
        venv_dir = BASE_DIR / venv_raw
    packages = cfg.get("packages")
    ensure_sandbox_venv(venv_dir, packages=packages)

    _sandbox_registry = SandboxRegistry(
        artifacts_dir=ARTIFACTS_DIR,
        venv_dir=venv_dir,
        runner_path=Path(__file__).parent / "sandbox_runner.py",
        timeout_seconds=int(cfg.get("timeout_seconds", 30)),
        max_output_chars=int(cfg.get("max_output_chars", 10_000)),
        idle_kill_minutes=int(cfg.get("idle_kill_minutes", 10)),
    )
    if bwrap_path() is None:
        print(
            "[sandbox] WARNING: bwrap is not installed. "
            "Python execution will run with resource limits only (no filesystem "
            "or network jail). Install `bubblewrap` for full sandboxing."
        )
    return _sandbox_registry


@app.on_event("startup")
async def _sandbox_startup():
    import asyncio
    global _sandbox_reaper_task

    async def reaper():
        while True:
            await asyncio.sleep(60)
            if _sandbox_registry is not None:
                killed = _sandbox_registry.reap_idle()
                if killed:
                    print(f"[sandbox] reaped idle sessions: {killed}")

    _sandbox_reaper_task = asyncio.create_task(reaper())


@app.on_event("shutdown")
async def _sandbox_shutdown():
    if _sandbox_reaper_task is not None:
        _sandbox_reaper_task.cancel()
    if _sandbox_registry is not None:
        _sandbox_registry.shutdown_all()


def load_config() -> dict:
    with open(CONFIG_PATH, "r") as f:
        return json.load(f)


# ── Models ─────────────────────────────────────────────────────────────────────

# Keywords that identify embedding models — excluded from the chat dropdown.
_EMBED_KEYWORDS = {"embed", "bge-m3", "nomic-embed"}


@app.get("/models")
async def get_models():
    """Query the local Ollama daemon for available models.

    Returns models split into two groups:
      - local:  models with actual weights on disk (size > 0 bytes)
      - cloud:  zero-byte placeholder models that proxy to a cloud service
    Embedding models are filtered out entirely.
    """
    config = load_config()
    ollama_url = os.getenv("OLLAMA_BASE_URL", config.get("ollama_base_url", "http://localhost:11434"))

    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            resp = await client.get(f"{ollama_url}/api/tags")
            resp.raise_for_status()
            data = resp.json()
    except Exception as exc:
        return {
            "models": [],
            "default_model": None,
            "error": f"Cannot reach Ollama at {ollama_url}: {exc}",
        }

    local_models: list[dict] = []
    cloud_models: list[dict] = []

    for m in data.get("models", []):
        name: str = m.get("name", "")
        size: int = m.get("size", 0)

        # Skip embedding models
        if any(kw in name.lower() for kw in _EMBED_KEYWORDS):
            continue

        # Models with actual local weights are at least tens of MB.
        # Cloud-routing stubs are just a Modelfile (< 1 KB).
        _LOCAL_THRESHOLD_BYTES = 50_000_000  # 50 MB

        is_local = size >= _LOCAL_THRESHOLD_BYTES
        entry = {
            "id": name,
            "name": name,
            "size_gb": round(size / 1_000_000_000, 1),
            "group": "local" if is_local else "cloud",
        }

        if is_local:
            local_models.append(entry)
        else:
            cloud_models.append(entry)

    local_models.sort(key=lambda x: x["name"])
    cloud_models.sort(key=lambda x: x["name"])

    all_models = local_models + cloud_models

    # Prefer the configured default if it exists in the list
    configured_default = config.get("default_model", "")
    if configured_default and any(m["id"] == configured_default for m in all_models):
        default = configured_default
    elif local_models:
        default = local_models[0]["id"]
    else:
        default = all_models[0]["id"] if all_models else None

    return {"models": all_models, "default_model": default}


# ── Backends ───────────────────────────────────────────────────────────────────

@app.get("/backends")
async def get_backends():
    """Get available agent backends."""
    config = load_config()
    return {
        "backends": agent_registry.list_backends(),
        "default_backend": config.get("default_backend", "react_agent"),
    }


# ── Model info / capabilities ──────────────────────────────────────────────────

@app.get("/model-info/{model_id:path}")
async def get_model_info(model_id: str):
    """Return capabilities for a specific model by querying Ollama /api/show.

    The 'thinking' capability is present for models that natively support
    reasoning mode (e.g. Qwen3, DeepSeek-R1, Gemma 4 IT with thinking, Nemotron).
    This is the authoritative source — works automatically with any new Ollama release.
    """
    config = load_config()
    ollama_url = os.getenv("OLLAMA_BASE_URL", config.get("ollama_base_url", "http://localhost:11434"))

    try:
        async with httpx.AsyncClient(timeout=8.0) as client:
            resp = await client.post(
                f"{ollama_url}/api/show",
                json={"name": model_id},
            )
            resp.raise_for_status()
            data = resp.json()
    except Exception as exc:
        return {
            "model_id": model_id,
            "supports_thinking": False,
            "capabilities": [],
            "error": str(exc),
        }

    capabilities: list[str] = data.get("capabilities", [])

    # Extract native context length (key varies: "llama.context_length", "qwen3.context_length", …)
    model_info = data.get("model_info", {})
    context_length: int | None = None
    for k, v in model_info.items():
        if "context_length" in k:
            context_length = int(v)
            break

    return {
        "model_id": model_id,
        "supports_thinking": "thinking" in capabilities,
        "supports_vision":   "vision"   in capabilities,
        "supports_tools":    "tools"    in capabilities,
        "capabilities": capabilities,
        "context_length": context_length,
    }


# ── Chat ───────────────────────────────────────────────────────────────────────

class ChatRequest(BaseModel):
    message: str = ""           # display text shown in UI (never contains base64 images)
    model_id: str
    conversation_id: Optional[str] = None
    web_search_enabled: bool = False
    num_search_results: int = 5
    backend_id: str = "react_agent"
    backend_config: dict = {}
    ollama_options: dict = {}   # num_predict, num_ctx, temperature, reasoning
    allow_network: bool = False # Python sandbox network access (frontend toggle)
    # Multimodal content blocks (text + image_url).  When set, used for storage
    # and sent to the model instead of `message`.  `message` is kept only as the
    # display label for history titles / PDF export header.
    message_content: Optional[list] = None


@app.post("/chat")
async def chat(request: ChatRequest):
    """Handle chat request with streaming SSE response."""
    conversation_id = request.conversation_id or str(uuid.uuid4())

    if conversation_id not in conversations:
        conversations[conversation_id] = {
            "messages": [],
            "model_id": request.model_id,
            "backend_id": request.backend_id,
        }

    # Use multimodal content when provided, otherwise plain text
    user_content = request.message_content if request.message_content else request.message
    conversations[conversation_id]["messages"].append(
        {"role": "user", "content": user_content}
    )
    conversations[conversation_id]["model_id"] = request.model_id
    conversations[conversation_id]["backend_id"] = request.backend_id

    async def generate_response() -> AsyncGenerator[str, None]:
        full_response = ""
        try:
            backend = agent_registry.get_backend(request.backend_id)
        except KeyError as e:
            yield f"data: {json.dumps({'type': 'error', 'error': str(e)})}\n\n"
            return

        config = load_config()
        code_exec_cfg = config.get("code_execution", {}) or {}
        registry = _get_sandbox_registry()
        tools = make_tools(
            ARTIFACTS_DIR,
            request.web_search_enabled,
            conversation_id=conversation_id,
            sandbox_registry=registry,
            allow_network=bool(request.allow_network),
            code_exec_config=code_exec_cfg,
        )

        # Resolve recursion_limit from config if not overridden
        merged_config = dict(request.backend_config)
        if "recursion_limit" not in merged_config and "recursion_limit" in config:
            merged_config["recursion_limit"] = config["recursion_limit"]

        # Backend-specific config sections (e.g. config["deep_agent"]) are merged
        # in so the backend can read its own knobs without re-loading the file.
        if request.backend_id in config and isinstance(config[request.backend_id], dict):
            for k, v in config[request.backend_id].items():
                merged_config.setdefault(k, v)

        # Pass ollama generation options through to the agent
        merged_config["ollama_options"] = request.ollama_options

        messages = [
            {"role": msg["role"], "content": msg["content"]}
            for msg in conversations[conversation_id]["messages"]
        ]

        try:
            async for event in backend.stream(
                messages=messages,
                model_id=request.model_id,
                tools=tools,
                backend_config=merged_config,
                conversation_id=conversation_id,
            ):
                if event.get("type") == "content":
                    full_response += event.get("content", "")
                yield f"data: {json.dumps(event)}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'error': str(e)})}\n\n"
            return

        conversations[conversation_id]["messages"].append(
            {"role": "assistant", "content": full_response}
        )

    return StreamingResponse(
        generate_response(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive"},
    )


@app.get("/conversation/{conversation_id}/messages")
async def get_conversation_messages(conversation_id: str, truncate_images: bool = True):
    """Return the raw messages list stored for a conversation — i.e. the exact
    payload the backend sends to the LLM on the next turn. Intended for debugging.

    When ``truncate_images`` is true (default), any ``image_url`` blocks inside
    multimodal content have their base64 data replaced with a short placeholder
    so the download stays readable.
    """
    if conversation_id not in conversations:
        raise HTTPException(status_code=404, detail="conversation not found")

    conv = conversations[conversation_id]

    def _sanitize_content(content):
        if not isinstance(content, list):
            return content
        out = []
        for block in content:
            if isinstance(block, dict) and block.get("type") == "image_url" and truncate_images:
                url = ""
                iu = block.get("image_url")
                if isinstance(iu, dict):
                    url = iu.get("url", "") or ""
                elif isinstance(iu, str):
                    url = iu
                prefix = url.split(",", 1)[0] if "," in url else url[:40]
                out.append({
                    "type": "image_url",
                    "image_url": {"url": f"{prefix},[base64 truncated, {len(url)} chars]"},
                })
            else:
                out.append(block)
        return out

    sanitized = [
        {"role": m.get("role"), "content": _sanitize_content(m.get("content"))}
        for m in conv.get("messages", [])
    ]

    return {
        "conversation_id": conversation_id,
        "model_id": conv.get("model_id"),
        "backend_id": conv.get("backend_id"),
        "message_count": len(sanitized),
        "messages": sanitized,
    }


# ── Spreadsheet helpers ────────────────────────────────────────────────────────

_EXCEL_EXTS = {".xlsx", ".xls", ".csv"}


def _fmt_cell(value) -> object:
    """Normalise a cell value: round floats to int, pass strings through, None stays None."""
    import math as _math
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if _math.isnan(value) or _math.isinf(value):
            return None
        return int(round(value))
    if isinstance(value, int):
        return value
    # datetime, date, time, and anything else → string
    return str(value)


def _sheet_to_json(columns: list, rows: list) -> str:
    """Serialise one sheet as compact JSON: {"columns":[...],"rows":[[...], ...]}."""
    data = {
        "columns": [str(c) if c is not None else "" for c in columns],
        "rows": rows,
    }
    return json.dumps(data, separators=(",", ":"), ensure_ascii=False)


def _read_xlsx_sheets(file_path, _source=None) -> list[tuple]:
    """Read an xlsx workbook. Pass _source=BytesIO to skip openpyxl's extension check."""
    import openpyxl, io as _io
    source = _source if _source is not None else str(file_path)
    wb = openpyxl.load_workbook(source, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        raw = list(ws.values)
        if not raw:
            sheets.append((name, [], []))
            continue
        columns = list(raw[0])
        rows = [
            [_fmt_cell(c) for c in row]
            for row in raw[1:]
            if any(c is not None for c in row)   # skip blank rows
        ]
        sheets.append((name, columns, rows))
    return sheets


def _read_xls_sheets(file_path) -> list[tuple]:
    import xlrd
    wb = xlrd.open_workbook(str(file_path))
    sheets = []
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        if ws.nrows == 0:
            sheets.append((name, [], []))
            continue
        columns = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
        rows = []
        for r in range(1, ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                ctype = cell.ctype
                if ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                elif ctype == xlrd.XL_CELL_NUMBER:
                    row.append(int(round(cell.value)))
                elif ctype == xlrd.XL_CELL_BOOLEAN:
                    row.append(bool(cell.value))
                else:
                    row.append(str(cell.value))
            if any(v is not None for v in row):
                rows.append(row)
        sheets.append((name, columns, rows))
    return sheets


def _read_csv_sheets(file_path) -> list[tuple]:
    import csv as _csv
    rows_raw = []
    # Try UTF-8 with BOM first (common in Excel-exported CSVs), fall back to latin-1
    for enc in ("utf-8-sig", "latin-1"):
        try:
            with open(file_path, newline="", encoding=enc) as f:
                rows_raw = list(_csv.reader(f))
            break
        except UnicodeDecodeError:
            continue
    if not rows_raw:
        return [("Sheet1", [], [])]
    columns = rows_raw[0]
    rows = []
    for raw_row in rows_raw[1:]:
        row = []
        for val in raw_row:
            s = val.strip()
            if not s:
                row.append(None)
            else:
                # Strip thousand-separator commas before attempting numeric parse
                try:
                    row.append(int(round(float(s.replace(",", "")))))
                except ValueError:
                    row.append(s)
        if any(v is not None for v in row):
            rows.append(row)
    return [("Sheet1", columns, rows)]


def _read_spreadsheet(file_path, filename: str) -> str:
    """Convert an Excel/CSV file to the tagged text format sent to the LLM."""
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        sheets = _read_csv_sheets(file_path)
    elif ext == ".xlsx":
        sheets = _read_xlsx_sheets(file_path)
    elif ext == ".xls":
        # Some files carry a .xls extension but are actually xlsx (XML/zip-based),
        # e.g. when saved from modern Excel with "Save as xls". Try the legacy BIFF
        # reader first; on failure read the raw bytes and pass them via BytesIO so
        # openpyxl skips its extension check and parses by content.
        try:
            sheets = _read_xls_sheets(file_path)
        except Exception:
            import io as _io
            with open(file_path, "rb") as _f:
                raw = _io.BytesIO(_f.read())
            sheets = _read_xlsx_sheets(file_path, _source=raw)
    else:
        raise ValueError(f"Unsupported spreadsheet extension: {ext}")

    parts = []
    for sheet_name, columns, rows in sheets:
        body = _sheet_to_json(columns, rows)
        parts.append(f'<sheet name="{sheet_name}">\n{body}\n</sheet>')
    return "\n".join(parts)


# ── Upload ─────────────────────────────────────────────────────────────────────

_ALLOWED_UPLOAD_EXTS = {".pdf"} | _EXCEL_EXTS


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload a PDF, Excel (.xlsx/.xls), or CSV file for use in the conversation."""
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in _ALLOWED_UPLOAD_EXTS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Allowed: PDF, XLSX, XLS, CSV.",
        )

    try:
        file_id = str(uuid.uuid4())
        file_path = UPLOADS_DIR / f"{file_id}{ext}"

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        # ── PDF ────────────────────────────────────────────────���──────────────
        if ext == ".pdf":
            text_content = ""
            images_base64 = []

            try:
                from pypdf import PdfReader
                reader = PdfReader(file_path)
                for page in reader.pages:
                    text_content += page.extract_text() or ""
                    text_content += "\n\n"
            except Exception as e:
                text_content = f"Error extracting text: {str(e)}"

            try:
                from pdf2image import convert_from_path
                import io
                images = convert_from_path(file_path, dpi=150)
                for i, img in enumerate(images):
                    img_buffer = io.BytesIO()
                    img.save(img_buffer, format="PNG")
                    img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
                    images_base64.append({"page": i + 1, "data": img_base64})
            except Exception:
                pass

            return {
                "file_id": file_id,
                "filename": file.filename,
                "file_type": "pdf",
                "text_content": text_content.strip(),
                "images": images_base64,
                "page_count": len(images_base64) if images_base64 else 0,
            }

        # ── Excel / CSV ───────────────────────────────────────────────────────
        else:
            text_content = _read_spreadsheet(file_path, file.filename)
            sheet_count = text_content.count('<sheet name=')
            return {
                "file_id": file_id,
                "filename": file.filename,
                "file_type": "excel",
                "text_content": text_content,
                "images": [],
                "page_count": sheet_count,
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── PDF export ─────────────────────────────────────────────────────────────────

class DownloadRequest(BaseModel):
    conversation_id: str


def _extract_user_display_text(raw) -> str:
    """Return a compact, human-readable string for a user message in the export.

    Strips the injected PDF context (the full extracted text wrapped in
    <pdf-filename-...> tags plus the numbered [Image N = ...] preamble lines)
    so the downloaded report is not bloated with the original PDF content.
    Replaces the block with a single '[PDF attached: filename.pdf]' note.
    Works for both plain strings and multimodal content lists.
    """
    if isinstance(raw, list):
        # Collect all text blocks; ignore image_url blocks entirely.
        text = "\n".join(
            b.get("text", "") for b in raw
            if isinstance(b, dict) and b.get("type") == "text"
        )
    else:
        text = raw or ""

    # Pull out the filenames mentioned in <pdf-filename-...> tags before stripping.
    filenames = re.findall(r"<pdf-filename-([^>]+)>", text)

    # Strip the entire <pdf-filename-FILENAME>...</pdf-filename-FILENAME> blocks.
    text = re.sub(r"<pdf-filename-[^>]+>[\s\S]*?</pdf-filename-[^>]+>", "", text)

    # Strip numbered image-reference preamble lines: [Image N = Page N of file.pdf]
    text = re.sub(r"^\[Image \d+ = Page \d+ of [^\]]+\]\n?", "", text, flags=re.MULTILINE)

    # Strip the "User question:" label that wraps the actual query.
    text = re.sub(r"^User question:\s*", "", text.strip())

    text = text.strip()

    # Prepend a compact PDF note so the reader knows a file was in context.
    if filenames:
        # Filenames were sanitised (spaces → underscores) when stored; restore for display.
        display_names = ", ".join(f.replace("_", " ") for f in filenames)
        text = f"[PDF attached: {display_names}]\n{text}" if text else f"[PDF attached: {display_names}]"

    return text


# ── Palette ───────────────────────────────────────────────────────────────────
_C_TITLE      = colors.HexColor("#0F172A")   # slate-900  — title, thick rule
_C_HEADING    = colors.HexColor("#1E293B")   # slate-800  — H1 / H2
_C_BODY       = colors.HexColor("#374151")   # gray-700   — body text (softer than pure black)
_C_META       = colors.HexColor("#9CA3AF")   # gray-400   — metadata, footer
_C_RULE       = colors.HexColor("#E5E7EB")   # gray-200   — turn dividers
_C_USER       = colors.HexColor("#1D4ED8")   # blue-700
_C_ASST       = colors.HexColor("#0F766E")   # teal-700
_C_TBL_HDR    = colors.HexColor("#F1F5F9")   # slate-100  — table header bg
_C_TBL_ALT    = colors.HexColor("#F8FAFC")   # slate-50   — alternating row bg
_C_TBL_BORDER = colors.HexColor("#CBD5E1")   # slate-300  — table grid lines

# Page geometry (Letter: 8.5 × 11 in)
_MARGIN_H = inch          # left / right
_MARGIN_T = 0.9 * inch    # top
_MARGIN_B = 0.75 * inch   # bottom  (footer lives below this)
_TEXT_W   = letter[0] - 2 * _MARGIN_H   # usable text width ≈ 6.5 in


def _escape_xml(text: str) -> str:
    """Escape &, < and > so ReportLab's XML parser doesn't choke."""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _md_inline(text: str) -> str:
    """Convert inline **bold** / *italic* markdown to ReportLab XML tags.
    Must be called AFTER _escape_xml so the inserted tags are not re-escaped.
    """
    text = re.sub(r"\*\*([^*\n]+?)\*\*", r"<b>\1</b>", text)
    text = re.sub(r"__([^_\n]+?)__",      r"<b>\1</b>", text)
    text = re.sub(r"(?<!\w)\*([^*\n]+?)\*(?!\w)", r"<i>\1</i>", text)
    text = re.sub(r"(?<!\w)_([^_\n]+?)_(?!\w)",   r"<i>\1</i>", text)
    return text


def _render_markdown_table(table_lines: list[str], tbl_body_style: ParagraphStyle) -> Table | None:
    """Convert a list of raw markdown table lines into a ReportLab Table flowable.

    Expects the standard GFM format:
        | Header | Header |
        |--------|--------|
        | Cell   | Cell   |
    The separator row (---|---) is detected and skipped; the first data row
    becomes the header row (bold).
    """
    _sep_re = re.compile(r"^\|[\s\-:|]+\|$")

    tbl_header_style = ParagraphStyle(
        "TblHdr", parent=tbl_body_style,
        fontName="Helvetica-Bold", fontSize=8.5, leading=12,
    )
    tbl_cell_style = ParagraphStyle(
        "TblCell", parent=tbl_body_style,
        fontName="Helvetica", fontSize=8.5, leading=12,
    )

    rows: list[list] = []
    is_first_data_row = True
    for line in table_lines:
        stripped = line.strip()
        if _sep_re.match(stripped):
            continue  # skip |---|---| separator
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        style = tbl_header_style if is_first_data_row else tbl_cell_style
        rows.append([
            Paragraph(_md_inline(_escape_xml(c)), style) for c in cells
        ])
        is_first_data_row = False

    if not rows:
        return None

    col_count = max(len(r) for r in rows)
    # Equalise column widths across the full text area
    col_w = _TEXT_W / col_count

    table = Table(rows, colWidths=[col_w] * col_count, hAlign="LEFT")
    table.setStyle(TableStyle([
        # Header row background
        ("BACKGROUND",    (0, 0), (-1, 0),  _C_TBL_HDR),
        # Alternating body rows
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, _C_TBL_ALT]),
        # Outer border + inner grid
        ("BOX",           (0, 0), (-1, -1), 0.5, _C_TBL_BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, _C_TBL_BORDER),
        # Slightly heavier rule under header
        ("LINEBELOW",     (0, 0), (-1, 0),  1,   _C_TBL_BORDER),
        # Cell padding
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 7),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
    ]))
    return table


def _format_content_for_pdf(
    content: str,
    body_style: ParagraphStyle,
    h1_style: ParagraphStyle,
    h2_style: ParagraphStyle,
    h3_style: ParagraphStyle,
    bullet_style: ParagraphStyle,
) -> list:
    """Parse markdown content into ReportLab flowables.

    Handles: H1/H2/H3, bullet lists, horizontal rules, GFM tables,
    and inline bold/italic in every element type.
    """
    elements = []
    content = content.replace("\r\n", "\n").replace("<br/>", "\n").replace("<br>", "\n")
    lines = content.split("\n")
    _sep_re = re.compile(r"^\|[\s\-:|]+\|$")

    i = 0
    while i < len(lines):
        s = lines[i].strip()

        # ── Markdown table detection ──────────────────────────────────────────
        # A table starts with a | row immediately followed by a |---|---| row.
        if s.startswith("|") and i + 1 < len(lines) and _sep_re.match(lines[i + 1].strip()):
            table_lines = []
            while i < len(lines) and "|" in lines[i]:
                table_lines.append(lines[i])
                i += 1
            tbl = _render_markdown_table(table_lines, body_style)
            if tbl:
                elements.append(Spacer(1, 6))
                elements.append(tbl)
                elements.append(Spacer(1, 8))
            continue

        # ── Standard block elements ───────────────────────────────────────────
        if not s:
            elements.append(Spacer(1, 3))

        elif s.startswith("### "):
            text = _md_inline(_escape_xml(s[4:].strip()))
            elements.append(Paragraph(f"<b>{text}</b>", h3_style))

        elif s.startswith("## "):
            text = _md_inline(_escape_xml(s[3:].strip()))
            elements.append(Paragraph(f"<b>{text}</b>", h2_style))

        elif s.startswith("# "):
            text = _md_inline(_escape_xml(s[2:].strip()))
            elements.append(Paragraph(f"<b>{text}</b>", h1_style))

        elif s in ("---", "***", "___"):
            elements.append(HRFlowable(
                width="100%", thickness=0.5, color=_C_RULE,
                spaceBefore=4, spaceAfter=4,
            ))

        elif re.match(r"^[-*\u2022]\s", s):
            bullet_text = _md_inline(_escape_xml(s[2:].strip()))
            elements.append(Paragraph(f"\u2022\u00a0{bullet_text}", bullet_style))

        else:
            text = _md_inline(_escape_xml(s))
            elements.append(Paragraph(text, body_style))

        i += 1

    return elements


# ── Numbered canvas — supports "Page X of Y" in footer ───────────────────────

class _NumberedCanvas(_rl_canvas.Canvas):
    """Two-pass canvas that stamps 'Page X of Y' once total page count is known."""

    def __init__(self, *args, **kwargs):
        _rl_canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_states: list[dict] = []
        self._model_label: str = ""

    def showPage(self):
        self._saved_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved_states)
        for state in self._saved_states:
            self.__dict__.update(state)
            self._stamp_chrome(total)
            _rl_canvas.Canvas.showPage(self)
        _rl_canvas.Canvas.save(self)

    def _stamp_chrome(self, total: int):
        page = self._pageNumber
        self.saveState()
        self.setFont("Helvetica", 7.5)
        self.setFillColor(_C_META)

        # Footer: right-aligned page counter
        self.drawRightString(
            letter[0] - _MARGIN_H,
            0.38 * inch,
            f"Page {page} of {total}",
        )
        # Footer: left-aligned model label (if provided)
        if self._model_label:
            self.drawString(_MARGIN_H, 0.38 * inch, self._model_label)

        # Running header on pages 2+ (thin rule + document title)
        if page > 1:
            self.setStrokeColor(_C_RULE)
            self.setLineWidth(0.5)
            y_rule = letter[1] - 0.52 * inch
            self.line(_MARGIN_H, y_rule, letter[0] - _MARGIN_H, y_rule)
            self.drawString(_MARGIN_H, letter[1] - 0.42 * inch, "Conversation Export")

        self.restoreState()


@app.post("/download-pdf")
async def download_conversation_pdf(request: DownloadRequest):
    """Download conversation as a polished, print-ready PDF."""
    conversation_id = request.conversation_id
    if conversation_id not in conversations:
        raise HTTPException(status_code=404, detail="Conversation not found")

    conversation = conversations[conversation_id]
    model_id     = conversation.get("model_id", "Unknown")
    pdf_filename  = (
        f"conversation_{conversation_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    )
    pdf_path = EXPORTS_DIR / pdf_filename

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=_MARGIN_H,
        rightMargin=_MARGIN_H,
        topMargin=_MARGIN_T,
        bottomMargin=_MARGIN_B,
    )

    # ── Paragraph styles (professional print scale) ───────────────────────────
    body_style = ParagraphStyle(
        "ExBody", fontName="Helvetica", fontSize=9.5, leading=14,
        spaceAfter=3, textColor=_C_BODY,
    )
    h1_style = ParagraphStyle(
        "ExH1", fontName="Helvetica-Bold", fontSize=13, leading=17,
        spaceBefore=12, spaceAfter=5, textColor=_C_HEADING,
    )
    h2_style = ParagraphStyle(
        "ExH2", fontName="Helvetica-Bold", fontSize=11, leading=15,
        spaceBefore=10, spaceAfter=3, textColor=_C_HEADING,
    )
    h3_style = ParagraphStyle(
        "ExH3", fontName="Helvetica-Bold", fontSize=10, leading=14,
        spaceBefore=8, spaceAfter=2, textColor=_C_BODY,
    )
    bullet_style = ParagraphStyle(
        "ExBullet", fontName="Helvetica", fontSize=9.5, leading=14,
        leftIndent=14, firstLineIndent=0, spaceAfter=2, textColor=_C_BODY,
    )
    meta_style = ParagraphStyle(
        "ExMeta", fontName="Helvetica", fontSize=8, leading=12,
        textColor=_C_META, spaceAfter=1,
    )
    title_style = ParagraphStyle(
        "ExTitle", fontName="Helvetica-Bold", fontSize=20, leading=25,
        spaceAfter=6, textColor=_C_TITLE,
    )
    user_label_style = ParagraphStyle(
        "ExUserLabel", fontName="Helvetica-Bold", fontSize=7, leading=10,
        spaceBefore=18, spaceAfter=4, textColor=_C_USER,
        tracking=60,   # letter-spacing for the small caps feel
    )
    asst_label_style = ParagraphStyle(
        "ExAsstLabel", fontName="Helvetica-Bold", fontSize=7, leading=10,
        spaceBefore=18, spaceAfter=4, textColor=_C_ASST,
        tracking=60,
    )

    # ── Document header ───────────────────────────────────────────────────────
    story = [
        Paragraph("Conversation Export", title_style),
        HRFlowable(width="100%", thickness=1, color=_C_TITLE, spaceAfter=8),
        Paragraph(f"Model&#160;&#160;{model_id}", meta_style),
        Paragraph(f"Date&#160;&#160;&#160;{datetime.now().strftime('%Y-%m-%d %H:%M')}", meta_style),
        Spacer(1, 20),
    ]

    # ── Message turns ─────────────────────────────────────────────────────────
    for i, msg in enumerate(conversation["messages"]):
        role = msg["role"]
        raw  = msg["content"]

        # Thin rule between turns; not before the first
        if i > 0:
            story.append(HRFlowable(
                width="100%", thickness=0.5, color=_C_RULE,
                spaceBefore=8, spaceAfter=0,
            ))

        if role == "user":
            content_text = _extract_user_display_text(raw)
            story.append(Paragraph("USER", user_label_style))
        else:
            content_text = raw or ""
            story.append(Paragraph("ASSISTANT", asst_label_style))

        story.extend(_format_content_for_pdf(
            content_text, body_style, h1_style, h2_style, h3_style, bullet_style,
        ))

    # Build with the two-pass numbered canvas
    def _make_canvas(*args, **kwargs):
        c = _NumberedCanvas(*args, **kwargs)
        c._model_label = model_id
        return c

    doc.build(story, canvasmaker=_make_canvas)

    return FileResponse(
        path=str(pdf_path),
        filename=pdf_filename,
        media_type="application/pdf",
    )


# ── Conversation management ────────────────────────────────────────────────────

@app.get("/conversation/{conversation_id}")
async def get_conversation(conversation_id: str):
    if conversation_id not in conversations:
        raise HTTPException(status_code=404, detail="Conversation not found")
    return conversations[conversation_id]


@app.delete("/conversation/{conversation_id}")
async def delete_conversation(conversation_id: str):
    if conversation_id in conversations:
        del conversations[conversation_id]
    if _sandbox_registry is not None:
        _sandbox_registry.shutdown_conversation(conversation_id)
    return {"status": "deleted"}


# ── Artifacts ──────────────────────────────────────────────────────────────────

@app.get("/artifacts")
async def get_artifacts():
    files = list(ARTIFACTS_DIR.glob("*"))
    return {
        "artifacts": [
            {
                "name": f.name,
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            }
            for f in files
            if f.is_file()
        ]
    }


@app.get("/artifacts/{filename}")
async def get_artifact(filename: str):
    filepath = ARTIFACTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Artifact not found")
    with open(filepath, "r") as f:
        content = f.read()
    return {"filename": filename, "content": content}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
